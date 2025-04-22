import models.FileAccess as file
import openpyxl
from openpyxl.formula.translate import Translator

class JudgeFile:
    def __init__(self, access, data, selected_currency) -> None:
        self.access = access
        self._csv_data = data
        self.selected_currency = selected_currency

    def __del__(self) -> None:
        pass

    def MakeFile(self):
        # NUM_FORMAT = '#,##0.00;[red]"-"#,##0.00'
        if 'JPY' in self.selected_currency:
            NUM_FORMAT = '0.00;[red]"-"0.00'
        else:
            NUM_FORMAT = '0.0000;[red]"-"0.0000'

        wb = openpyxl.Workbook()
        sheet = wb.active
        # sheet.title = 'test_sheet_1'

        i = 1
        for val in ['Date', self.selected_currency, 'EMA12', 'EMA26', 'MACD', 'SIGNAL', 'JUDGE']:
            sheet.cell(1,i).value = val
            i += 1

        a_max_length = 0

        i = 2
        for d in self._csv_data:
            sheet.cell(i,1).value = d[0]
            sheet.cell(i,2).value = d[1]

            if i == 13:
                sheet.cell(i,3).value = '=AVERAGE(B2:B13)'
            elif i == 27:
                sheet.cell(i,4).value = '=AVERAGE(B2:B27)'
                sheet.cell(i,5).value = '=C27-D27'
            elif i == 35:
                sheet.cell(i,6).value = '=AVERAGE(E27:E35)'
                sheet.cell(i,7).value = '=E35-F35'

            if i >= 14:
                sheet[sheet.cell(i,3).coordinate]= Translator('=(B14*2+C13*11)/13', origin='C14').translate_formula(sheet.cell(i,3).coordinate)

            if i >= 28:
                sheet[sheet.cell(i,4).coordinate]= Translator('=(B28*2+D27*25)/27', origin='D28').translate_formula(sheet.cell(i,4).coordinate)
                sheet[sheet.cell(i,5).coordinate]= Translator('=C28-D28', origin='E28').translate_formula(sheet.cell(i,5).coordinate)

            if i >= 36:
                sheet[sheet.cell(i,6).coordinate]= Translator('=(E36*2+F35*8)/10', origin='F36').translate_formula(sheet.cell(i,6).coordinate)
                sheet[sheet.cell(i,7).coordinate]= Translator('=E36-F36', origin='G36').translate_formula(sheet.cell(i,7).coordinate)

            if len(str(sheet.cell(i,1).value)) > a_max_length:
                a_max_length = len(str(sheet.cell(i,1).value))

            sheet[sheet.cell(i,2).coordinate].number_format = NUM_FORMAT
            sheet[sheet.cell(i,3).coordinate].number_format = NUM_FORMAT
            sheet[sheet.cell(i,4).coordinate].number_format = NUM_FORMAT
            sheet[sheet.cell(i,5).coordinate].number_format = NUM_FORMAT
            sheet[sheet.cell(i,6).coordinate].number_format = NUM_FORMAT
            sheet[sheet.cell(i,7).coordinate].number_format = NUM_FORMAT

            i += 1
        
        sheet.column_dimensions['A'].width = (a_max_length + 2) * 1.2
        self.access.save_judge_file(wb)
