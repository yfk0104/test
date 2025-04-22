import openpyxl.utils
import models.FileAccess as file
import openpyxl
import numpy as np
import models.Define as define

class SecretChart:
    def __init__(self, access, data, selected_currency, interval) -> None:
        self._access = access
        self._csv_data = data
        self.selected_currency = selected_currency
        self._interval = interval
        self._sheet = None
        self.now_row = 0
        self.row_max = 0
        self.row_min = 0
        self.now_col = 0
        self.direction = define.Direction.START

    def __del__(self) -> None:
        pass

    def _WriteBase(self):
        rate_list = [row[1] for row in self._csv_data]
        max_data = max(rate_list) + (self._interval * 5)
        min_data = min(rate_list) - (self._interval * 5)

        if 'JPY' in self.selected_currency:
            digit = 2
        else:
            digit = 5

        i = 1
        for r in np.arange(max_data, min_data, -1.0 * self._interval):
            self._sheet.cell(i,1).value = round(r, digit)
            i += 1

    def _WritePoint(self, cel):
        if cel.row > self.now_row:       # レートが低くなった
            i = 0
            for i in range(cel.row - self.now_row):
                self._sheet.cell(cel.row - i, self.now_col).value = '〇'

            self.now_row = cel.row
            self.row_max = self.now_row
            self.direction = define.Direction.DOWN

        elif cel.row < self.now_row:                         # レートが高くなった
            i = 0
            for i in range(self.now_row - cel.row):
                self._sheet.cell(cel.row + i, self.now_col).value = '×'

            self.now_row = cel.row
            self.row_min = self.now_row
            self.direction = define.Direction.UP

        else:
            None

    def _Write(self):
        index = 0

        for day in self._csv_data:                  # 日ごとのレートを古いものから順番
            for cel in self._sheet['A']:            # A列を上から順番
                if cel.value < day[1]:              # 本日の枠を見つける
                    if self.direction == define.Direction.START:
                        cel = cel.offset(0,1)
                        cel.value = '■'
                        self.now_row = cel.row
                        self.now_col = cel.column
                        self.direction = define.Direction.INIT
                    elif self.direction == define.Direction.INIT:
                        self.now_col += 1
                        self._WritePoint(cel)
                    elif self.direction == define.Direction.UP:
                        if (cel.row - self.now_row) >= 3: # 反転しているか？
                            self.now_col += 1
                            self._WritePoint(cel)

                        if cel.row < self.row_min:
                            self._WritePoint(cel)
                    elif self.direction == define.Direction.DOWN:
                        if (self.now_row - cel.row) >= 3: # 反転しているか？
                            self.now_col += 1
                            self._WritePoint(cel)

                        if cel.row > self.row_max:
                            self._WritePoint(cel)

                    break #ポイントを書いたらbreakする

            index += 1

    def MakeFile(self):
        wb = openpyxl.Workbook()
        self._sheet = wb.active

        self._WriteBase()
        self._Write()

        for col in range(2, self.now_col + 1):
            # 行の幅を変更
            self._sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3

        self._access.save_chart_file(wb)
        
